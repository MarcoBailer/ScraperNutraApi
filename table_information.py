import g4f
from g4f.client import Client
import json
import os
from openpyxl import Workbook, load_workbook

client = Client()

local_image = open("image.png", "rb")
response_local = client.chat.completions.create(
    model=g4f.models.default_vision,
    messages=[
        {"role": "user", "content": "Leia conteudo da tabela INFORMAÇÃO NUTRICIONAL. Ignore conteúdo que não são da tabela. Coloque as informações no formato json, coloque apenas um json na resposta, nada mais. O valor padrão se não houver inteiro é 0.  "
        """"
{
  "alimento": {
    "descricao": "Nome do alimento",
    "marca": "Marca do alimento",
    "categoria": "Categoria do alimento",    
  },
  {
  "porcao": {
    "porcao_por_embalagem": ,
    "porcao": ,
  },
  "valores_por_100g": {
    "valor_energetico_kcal": double,
    "carboidratos_g": double,
    "acucares_totais_g": double,
    "acucares_adicionados_g": double,
    "proteinas_g": double,
    "gorduras_totais_g": double,
    "gorduras_saturadas_g": double,
    "gorduras_trans_g": double,
    "fibras_g": double,
    "sodio_mg": 16,
    "vitamina_a_mcg": double,
    "vitamina_c_mg": double,
    "vitamina_d_mcg": double,
    "vitamina_b12_mcg": double,
    "vitamina_e_mg": double,
    "Tiamina_mg": double,
    "riboflavina_mg": double,
    "niacina_mg": double,
    "vitamina_b6_mg": double,
    "acido_folico_mcg": double,
    "biotina_mcg": double,
    "acido_pantotenico_mg": double,
    "vitamina_k_mcg": double,
    "ferro_mg": double,
    "calcio_mg": double,
    "zinco_mg": double,
    "potassio_mg": double,
    "magnesio_mg": double,
    "iodo_mcg": double,
    "cobre_mg": double,
    "manganes_mg": double,
    "selenio_mcg": double
    "cromo_mcg": double,
    "molibdenio_mcg": double
  }
}
}
Demais informações colocar em observacao.
"""}
    ],
    image=local_image
)
print("Response for local image:")
print(response_local.choices[0].message.content)
local_image.close()


def salvar_em_excel(json_str, arquivo_excel='informacao_nutricional.xlsx'):
    try:
        dados = json.loads(json_str)
    except Exception as e:
        print("Erro ao carregar JSON:", e)
        return

    # Define os campos e a ordem das colunas
    colunas = [
        "descricao", "marca", "categoria",
        "porcao_por_embalagem", "porcao",
        "valor_energetico_kcal", "carboidratos_g", "acucares_totais_g", "acucares_adicionados_g",
        "proteinas_g", "gorduras_totais_g", "gorduras_saturadas_g", "gorduras_trans_g",
        "fibras_g", "sodio_mg", "vitamina_a_mcg", "vitamina_c_mg", "vitamina_d_mcg",
        "vitamina_b12_mcg","vitamina_e_mg", "Tiamina_mg", "riboflavina_mg", "niacina_mg",
        "vitamina_b6_mg", "acido_folico_mcg", "biotina_mcg", "acido_pantotenico_mg",
        "vitamina_k_mcg", "ferro_mg", "calcio_mg", "zinco_mg", "potassio_mg",
        "magnesio_mg", "iodo_mcg", "cobre_mg", "manganes_mg", "selenio_mcg",
        "cromo_mcg", "molibdenio_mcg"
    ]

    # Extrai os dados do JSON
    alimento = dados.get("alimento", {})
    porcao = dados.get("porcao", {})
    valores = dados.get("valores_por_100g", {})

    linha = [
        alimento.get("descricao", ""),
        alimento.get("marca", ""),
        alimento.get("categoria", ""),
        porcao.get("porcao_por_embalagem", 0),
        porcao.get("porcao", 0),
        valores.get("valor_energetico_kcal", 0),
        valores.get("carboidratos_g", 0),
        valores.get("acucares_totais_g", 0),
        valores.get("acucares_adicionados_g", 0),
        valores.get("proteinas_g", 0),
        valores.get("gorduras_totais_g", 0),
        valores.get("gorduras_saturadas_g", 0),
        valores.get("gorduras_trans_g", 0),
        valores.get("fibras_g", 0),
        valores.get("sodio_mg", 0),
        valores.get("vitamina_a_mcg", 0),
        valores.get("vitamina_c_mg", 0),
        valores.get("vitamina_d_mcg", 0),
        valores.get("vitamina_b12_mcg", 0),
        valores.get("vitamina_e_mg", 0),
        valores.get("Tiamina_mg", 0),
        valores.get("riboflavina_mg", 0),
        valores.get("niacina_mg", 0),
        valores.get("vitamina_b6_mg", 0),
        valores.get("acido_folico_mcg", 0),
        valores.get("biotina_mcg", 0),
        valores.get("acido_pantotenico_mg", 0),
        valores.get("vitamina_k_mcg", 0),
        valores.get("ferro_mg", 0),
        valores.get("calcio_mg", 0),
        valores.get("zinco_mg", 0),
        valores.get("potassio_mg", 0),
        valores.get("magnesio_mg", 0),
        valores.get("iodo_mcg", 0),
        valores.get("cobre_mg", 0),
        valores.get("manganes_mg", 0),
        valores.get("selenio_mcg", 0),
        valores.get("cromo_mcg", 0),
        valores.get("molibdenio_mcg", 0)        
    ]

    # Cria ou abre o arquivo Excel
    if os.path.exists(arquivo_excel):
        wb = load_workbook(arquivo_excel)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(colunas)

    ws.append(linha)
    wb.save(arquivo_excel)
    print(f"Dados salvos em {arquivo_excel}")

content = response_local.choices[0].message.content.strip()
if content.startswith("```"):
    content = "\n".join(line for line in content.splitlines() if not line.strip().startswith("```"))
salvar_em_excel(content)